import os
import sys
from openpyxl import Workbook
import argparse
import logging

from datetime import datetime
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
def parse_bullshit_js(lines):
    created_on_txt = ""
    tweet_txt = ""
    data = []
    for line in lines:

        if "\"created_at\" : " in line:
            start_of_date = line.index("\"created_at\" : ") + 16
            created_on_txt = line[start_of_date:-1]
        if "\"full_text\" : " in line:
            start_of_tweet = line.index("\"full_text\" : ") + 15
            tweet_txt = line[start_of_tweet:-1]
            data.append({"created_at": created_on_txt, "tweet_text": tweet_txt})
    return data
def export(json_file_in, excel_file_out):
    #set up our excel workbook/worksheet
    if os.path.exists(json_file_in):
        start_of_load = datetime.now()
        logger.info(f"Loading json file {json_file_in}")
        with open(json_file_in) as json_file:
            lines = json_file.readlines()
        twiter_data = parse_bullshit_js(lines)
        logger.info(f"Loaded bs js file {json_file_in}")
        logger.info("Starting export")
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        ws["A1"] = "Date"
        ws["B1"] = "Tweet"
        row = 2
        for tweet in twiter_data: #target the array with the tweets
            row_str = str(row)
            ws["A" + row_str] = tweet["created_at"]
            ws["B" + row_str] = tweet["tweet_text"]
            row += 1
        wb.save(excel_file_out)
    else:
        print(f"{json_file_in} doesn't exist")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(prog='twiter_archive_to_excel',
                    description='Read the json archive file from twiter and export the data to an xlsx file.')
    parser.add_argument("--js-file-in", required=True, help="The tweets js file you want to export from")
    parser.add_argument("--excel-file-out", required=True, help="The excel file you want save to")
    if len(sys.argv) == 1:
        parser.print_help()
        sys.exit(1)
    args = parser.parse_args()
    export(args.json_file_in, args.excel_file_out)
